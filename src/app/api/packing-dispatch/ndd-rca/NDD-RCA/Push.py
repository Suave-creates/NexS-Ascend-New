"""Push the RCA pipeline outputs into the Google Sheet.

  1. "Raw Power BI"    tab <- order_level_data.csv (produced by Fetch.py), pasted as-is
  2. "Packing Data"    tab <- PackingScan + FR0Scan rows from the warehouse DB
  3. "Dispatched AWBs" tab <- ndd_shipments AWBs for the RCA date (Normal+Rescue->A, Rescue->B)

Tabs are cleared (values only, cells/formatting kept) and rewritten; "Dispatched
AWBs" only touches columns A and B. Existing contents are backed up to
backup_<tab>.csv first.

Google auth: first run opens a browser for one-time consent (write scope), then
caches gsheet_token.json. Sign in with the account that can edit the sheet.

  python Push.py            # do all (DB steps skipped automatically if unreachable)
  python Push.py raw        # only the Raw Power BI tab
  python Push.py packing    # only the Packing Data tab
  python Push.py awb        # only the Dispatched AWBs tab
  python Push.py qcf        # QC-fail journey timestamps -> QCF Analysis tab (M,N,O,P,R,T,V,X,Z)
  python Push.py check      # auth + list tabs + back up, no clearing/writing

  pip install gspread google-auth-oauthlib mysql-connector-python pandas
"""
import sys
import os
import datetime as dt

import pandas as pd

# --- Config ---
SPREADSHEET_ID = "19HMoPcNGCMtBX1alH5lB1xT88UzsIPrXAGCwaDZZyxQ"
RAW_TAB  = "Raw Power BI"
PACK_TAB = "Packing Data"
AWB_TAB  = "Dispatched AWBs"
QCF_TAB  = "QCF Analysis"
CSV_PATH = "order_level_data.csv"

# RCA date drives everything; the Dashboard sets NDD_RCA_DATE. Override the
# default here only when running Push.py standalone for a different day.
RCA_DATE   = os.environ.get("NDD_RCA_DATE", "2026-06-12")    # the order date analysed
_d = pd.to_datetime(RCA_DATE).date()
WINDOW_LO  = f"{_d - pd.Timedelta(days=2)} 00:00:00"         # RCA date minus 2 days
WINDOW_HI  = f"{_d + pd.Timedelta(days=1)} 23:59:59"         # through RCA date + 1 day

# Tabs to freeze to static values and export as an .xlsx.
# (source tab, output sheet, key_col, keep_format, as_is):
#   key_col     — if set, trailing rows where that column is blank are dropped (CALC 2 has
#                 a 21k-row formula grid where only helper cols are filled below real data).
#   keep_format — pull cell background colours, bold and merges from the source. RAW BI is
#                 plain raw data (no colours/merges) so we skip it to keep the payload small.
#   as_is       — paste the source exactly (its own colours, fonts, alignment, wrap, merges,
#                 full width incl. blank spacer cols), then add thin borders ONLY on A1:R37.
#                 Non-as_is tabs get trimmed + centered + all-borders.
EXPORT_TABS = [
    ("RCA", "RCA", None, True, True),
    ("BI Segregated", "RAW BI", None, False, False),
    ("CALC 2", "Study", 0, True, False),
    # QCF Analysis is a 21k-row formula grid where only the ~310 rows with a fitting_id
    # in col G (index 6) are real; trim by that key, keep the report's colours/merges.
    ("QCF Analysis", "QCF Analysis", 6, True, False),
]
RCA_BORDER_ROWS, RCA_BORDER_COLS = 37, 18    # A1:R37

# Finished RCA workbooks now live on Google Drive (no longer written to disk):
#   <DRCA_DRIVE_FOLDER>/<Year>/<MonthName>/DRCA D-M-YYYY.xlsx
# Year and Month subfolders are created on demand; a same-named file is overwritten.
DRCA_DRIVE_FOLDER = "1Kygr8PWqaZ8t0GAZDnB3nBJIgriAeC9U"

def drca_name(date_str=None):
    d = pd.to_datetime(date_str or RCA_DATE).date()
    return f"DRCA {d.day}-{d.month}-{d.year}.xlsx"     # day/month NOT zero-padded

DB = dict(host="192.168.27.157", port=3306, user="Hero",
          password="Lenskart@123", database="mydb", connection_timeout=15)

PACKING_SQL = """
SELECT scanID AS `Scan ID`, stationID AS `Station ID`,
       timestamp AS `Time Stamp`, nexsId AS `NexS ID`
FROM PackingScan
WHERE timestamp BETWEEN %s AND %s
ORDER BY timestamp ASC
"""
FR0_SQL = """
SELECT scanID AS `Scan ID`, stationID AS `Station ID`,
       createdAt AS `Time Stamp`, nexsId AS `NexS ID`
FROM FR0Scan
WHERE createdAt BETWEEN %s AND %s
ORDER BY createdAt ASC
"""
# Dispatched AWBs: awb where the created_at date matches the RCA date; all dispatched
# (Normal+Rescue) -> col A, Rescue -> col B (only those two columns are touched; the
# rest of the tab is left alone).
AWB_SQL = """
SELECT awb, type
FROM ndd_shipments
WHERE DATE(created_at) = %s
ORDER BY id ASC
"""

# QC-fail journey timestamps for the QCF Analysis tab. wms is a SEPARATE server from
# the mydb warehouse box above (different host/port/creds). All times UTC -> IST.
# Two passes per fitting_id: journey 0 (qc_fail_count=0, pre-fail) and journey 1
# (qc_fail_count=1, the re-processing after the fail).
WMS_DB = dict(host="192.168.27.132", port=13307, user="AryaK",
              password="inHTeH!HokX2GRXZHs67H", database="wms", connection_timeout=15)

QCF_SQL = """
SELECT
  fitting_id,
  MIN(CASE WHEN operation='IN_QC'     AND qc_fail_count=0 THEN CONVERT_TZ(updated_at,'+00:00','+05:30') END) AS IN_QC_0,
  MIN(CASE WHEN operation='QC_HOLD'   AND qc_fail_count=0 THEN CONVERT_TZ(updated_at,'+00:00','+05:30') END) AS QC_HOLD_0,
  MIN(CASE WHEN operation='QC_UNHOLD' AND qc_fail_count=0 THEN CONVERT_TZ(updated_at,'+00:00','+05:30') END) AS QC_UNHOLD_0,
  MIN(CASE WHEN operation='QC_FAILED' AND qc_fail_count=0 THEN CONVERT_TZ(updated_at,'+00:00','+05:30') END) AS QC_FAILED_0,
  MAX(CASE WHEN operation='IN_TRAY'                AND qc_fail_count=1 THEN CONVERT_TZ(updated_at,'+00:00','+05:30') END) AS IN_TRAY_1,
  MAX(CASE WHEN operation='EDGING'                 AND qc_fail_count=1 THEN CONVERT_TZ(updated_at,'+00:00','+05:30') END) AS EDGING_1,
  MAX(CASE WHEN operation='PENDING_CUSTOMIZATION'  AND qc_fail_count=1 THEN CONVERT_TZ(updated_at,'+00:00','+05:30') END) AS PENDING_CUSTOMIZATION_1,
  MAX(CASE WHEN operation='CUSTOMIZATION_COMPLETE' AND qc_fail_count=1 THEN CONVERT_TZ(updated_at,'+00:00','+05:30') END) AS CUSTOMIZATION_COMPLETE_1,
  MAX(CASE WHEN operation='QC_DONE'                AND qc_fail_count=1 THEN CONVERT_TZ(updated_at,'+00:00','+05:30') END) AS QC_DONE_1
FROM wms.order_items_history
WHERE fitting_id IN ({placeholders})
GROUP BY fitting_id
"""
# The 9 SELECT value columns (after fitting_id) map 1:1, in order, to these QCF Analysis
# columns. Q (TIME SPENT in QC) and S (Time SPENT QCF PICKING) are left alone — they are
# sheet formulas the user maintains.
#   IN_QC_0->M  QC_HOLD_0->N  QC_UNHOLD_0->O  QC_FAILED_0->P  IN_TRAY_1->R
#   EDGING_1->T  PENDING_CUSTOMIZATION_1->V  CUSTOMIZATION_COMPLETE_1->X  QC_DONE_1->Z
QCF_COLS = ["M", "N", "O", "P", "R", "T", "V", "X", "Z"]

HERE = os.path.dirname(os.path.abspath(__file__))
RESOURCE_ROOT = os.path.abspath(os.path.join(HERE, "../../../../../..", "src", "utils", "resources"))
CREDS = os.path.join(RESOURCE_ROOT, "google", "gcreds.json")
TOKEN = os.path.join(RESOURCE_ROOT, "google", "gsheet_token.json")
# Drive scope is needed to upload the finished workbook. The original token was minted
# for Sheets only; adding the scope forces a one-time re-consent (a refresh never widens
# scope). After the next browser sign-in the new token carries both.
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def get_creds():
    """OAuth user creds with Sheets + Drive scope. Browser consent on first run or when
    the cached token was granted fewer scopes than we now need; silent refresh after."""
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request

    creds = None
    if os.path.exists(TOKEN):
        # Load WITHOUT forcing SCOPES so creds.scopes reflects what was actually GRANTED
        # (the original token was Sheets-only). If we passed SCOPES here, has_scopes()
        # would report the wider Drive scope as present and the refresh would then fail
        # with 'invalid_scope' — the refresh token can't be widened, only re-consented.
        creds = Credentials.from_authorized_user_file(TOKEN)

    if creds and creds.has_scopes(SCOPES):
        if creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                with open(TOKEN, "w") as f:
                    f.write(creds.to_json())
            except Exception:
                creds = None          # refresh token revoked/expired -> re-consent
    else:
        creds = None                  # missing or narrower-scope token -> re-consent

    if not creds or not creds.valid:
        creds = InstalledAppFlow.from_client_secrets_file(CREDS, SCOPES).run_local_server(port=0)
        with open(TOKEN, "w") as f:
            f.write(creds.to_json())
    return creds


def get_sheet():
    """gspread Spreadsheet handle; browser consent on first run, cached after."""
    import gspread
    return gspread.authorize(get_creds()).open_by_key(SPREADSHEET_ID)


def get_drive():
    """Google Drive v3 service, sharing the same OAuth creds as the Sheets client."""
    from googleapiclient.discovery import build
    return build("drive", "v3", credentials=get_creds())


def _drive_subfolder(svc, name, parent_id):
    """Return the id of folder `name` under `parent_id`, creating it if absent."""
    safe = str(name).replace("'", "\\'")
    q = (f"name = '{safe}' and mimeType = 'application/vnd.google-apps.folder' "
         f"and '{parent_id}' in parents and trashed = false")
    hits = svc.files().list(q=q, spaces="drive", fields="files(id)",
                            supportsAllDrives=True,
                            includeItemsFromAllDrives=True).execute().get("files", [])
    if hits:
        return hits[0]["id"]
    meta = {"name": str(name), "mimeType": "application/vnd.google-apps.folder",
            "parents": [parent_id]}
    return svc.files().create(body=meta, fields="id",
                              supportsAllDrives=True).execute()["id"]


def upload_drca_to_drive(data, fname, year, month):
    """Upload workbook bytes to <DRCA_DRIVE_FOLDER>/<year>/<month>/<fname>, overwriting an
    existing same-named file in place. Returns its webViewLink."""
    from googleapiclient.http import MediaInMemoryUpload
    svc = get_drive()
    year_id = _drive_subfolder(svc, year, DRCA_DRIVE_FOLDER)
    month_id = _drive_subfolder(svc, month, year_id)

    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    media = MediaInMemoryUpload(data, mimetype=mime, resumable=False)
    safe = fname.replace("'", "\\'")
    existing = svc.files().list(
        q=f"name = '{safe}' and '{month_id}' in parents and trashed = false",
        spaces="drive", fields="files(id)", supportsAllDrives=True,
        includeItemsFromAllDrives=True).execute().get("files", [])
    if existing:
        f = svc.files().update(fileId=existing[0]["id"], media_body=media,
                               fields="id,webViewLink",
                               supportsAllDrives=True).execute()
    else:
        f = svc.files().create(body={"name": fname, "parents": [month_id]},
                               media_body=media, fields="id,webViewLink",
                               supportsAllDrives=True).execute()
    return f.get("webViewLink", "")


def drca_drive_link(date_str=None):
    """webViewLink of the DRCA file for the date on Drive, or '' if it isn't there yet."""
    d = pd.to_datetime(date_str or RCA_DATE).date()
    svc = get_drive()

    def child_folder(name, parent):
        safe = str(name).replace("'", "\\'")
        hits = svc.files().list(
            q=(f"name = '{safe}' and mimeType = 'application/vnd.google-apps.folder' "
               f"and '{parent}' in parents and trashed = false"),
            spaces="drive", fields="files(id)", supportsAllDrives=True,
            includeItemsFromAllDrives=True).execute().get("files", [])
        return hits[0]["id"] if hits else None

    year_id = child_folder(d.year, DRCA_DRIVE_FOLDER)
    if not year_id:
        return ""
    month_id = child_folder(d.strftime("%B"), year_id)
    if not month_id:
        return ""
    safe = drca_name(date_str).replace("'", "\\'")
    hits = svc.files().list(
        q=f"name = '{safe}' and '{month_id}' in parents and trashed = false",
        spaces="drive", fields="files(id,webViewLink)", supportsAllDrives=True,
        includeItemsFromAllDrives=True).execute().get("files", [])
    return hits[0].get("webViewLink", "") if hits else ""


def write_tab(sh, title, df):
    """Back up, clear (values only), and write df (with header) starting at A1.
    Grid is only ever grown, never shrunk, so no existing cell is deleted."""
    ws = sh.worksheet(title)

    backup = ws.get_all_values()
    bpath = os.path.join(HERE, f"backup_{title.replace(' ', '_')}.csv")
    pd.DataFrame(backup).to_csv(bpath, index=False, header=False)
    print(f"  backed up {len(backup)} rows -> {os.path.basename(bpath)}")

    values = [list(df.columns)] + df.where(pd.notnull(df), "").astype(str).values.tolist()
    n_rows, n_cols = len(values), len(values[0])
    if ws.row_count < n_rows or ws.col_count < n_cols:
        ws.resize(rows=max(ws.row_count, n_rows), cols=max(ws.col_count, n_cols))
    ws.clear()

    # Chunk by cells, not rows: Sheets allows 60 write requests/min/user, so keep
    # the request count low (~100k cells/request) regardless of column count.
    chunk = max(1, 100_000 // n_cols)
    for i in range(0, n_rows, chunk):
        _update_with_retry(ws, f"A{i + 1}", values[i:i + chunk])
    print(f"  wrote {n_rows - 1} data rows x {n_cols} cols to '{title}'")


def _update_with_retry(ws, rng, block, value_input_option="USER_ENTERED"):
    """One values.update with exponential backoff on the 429 write-quota error."""
    import time
    from gspread.exceptions import APIError
    delay = 5
    for attempt in range(6):
        try:
            ws.update(range_name=rng, values=block, value_input_option=value_input_option)
            time.sleep(1.2)   # gentle pacing between requests
            return
        except APIError as e:
            if "429" in str(e) and attempt < 5:
                print(f"    429 quota hit, backing off {delay}s...")
                time.sleep(delay)
                delay *= 2
                continue
            raise


def fetch_packing_df():
    import pymysql                       # pure-Python; mysql-connector C-ext segfaults here
    cfg = dict(DB)
    cfg["connect_timeout"] = cfg.pop("connection_timeout", 15)
    cn = pymysql.connect(**cfg)
    try:
        frames = []
        with cn.cursor() as cur:
            for label, sql in (("PackingScan", PACKING_SQL), ("FR0Scan", FR0_SQL)):
                cur.execute(sql, (WINDOW_LO, WINDOW_HI))
                cols = [d[0] for d in cur.description]
                rows = cur.fetchall()
                frames.append(pd.DataFrame(rows, columns=cols))
                print(f"  {label}: {len(rows)} rows")
    finally:
        cn.close()
    return pd.concat(frames, ignore_index=True)   # FR0 appended below Packing


def fetch_awbs():
    """AWBs dispatched on the RCA date, split into (normal, rescue) lists."""
    import pymysql
    cfg = dict(DB)
    cfg["connect_timeout"] = cfg.pop("connection_timeout", 15)
    cn = pymysql.connect(**cfg)
    try:
        with cn.cursor() as cur:
            cur.execute(AWB_SQL, (RCA_DATE,))
            rows = cur.fetchall()
    finally:
        cn.close()
    normal = [str(awb) for awb, typ in rows if typ == "Normal"]
    rescue = [str(awb) for awb, typ in rows if typ == "Rescue"]
    other = sorted({typ for _, typ in rows} - {"Normal", "Rescue"})
    print(f"  ndd_shipments {RCA_DATE}: Normal={len(normal)} Rescue={len(rescue)}"
          + (f"  (ignored types: {other})" if other else ""))
    return normal, rescue


def write_awbs(sh, normal, rescue):
    """Replace ONLY columns A and B of the AWB tab; leave C+ alone. Column A holds all
    dispatched AWBs (Normal followed by Rescue); column B holds Rescue only.
    AWBs are written RAW so 15-digit numeric AWBs stay exact text, not scientific."""
    ws = sh.worksheet(AWB_TAB)

    current = ws.get_all_values()
    bpath = os.path.join(HERE, f"backup_{AWB_TAB.replace(' ', '_')}.csv")
    pd.DataFrame([r[:2] for r in current]).to_csv(bpath, index=False, header=False)
    print(f"  backed up cols A:B of {len(current)} rows -> {os.path.basename(bpath)}")

    dispatched = normal + rescue                  # col A = every dispatched AWB
    if len(dispatched) > ws.row_count or len(rescue) > ws.row_count:
        ws.resize(rows=max(ws.row_count, len(dispatched), len(rescue)), cols=ws.col_count)
    ws.batch_clear(["A:B"])                       # clear old AWBs in A and B only
    if dispatched:
        _update_with_retry(ws, "A1", [[a] for a in dispatched], value_input_option="RAW")
    if rescue:
        _update_with_retry(ws, "B1", [[r] for r in rescue], value_input_option="RAW")
    print(f"  wrote Dispatched->A ({len(dispatched)}: Normal {len(normal)}+Rescue {len(rescue)}), "
          f"Rescue->B ({len(rescue)}) in '{AWB_TAB}'")


def _batch_update_with_retry(ws, batch, value_input_option="USER_ENTERED"):
    """ws.batch_update of several {range,values} blocks, with 429 backoff."""
    import time
    from gspread.exceptions import APIError
    delay = 5
    for attempt in range(6):
        try:
            ws.batch_update(batch, value_input_option=value_input_option)
            return
        except APIError as e:
            if "429" in str(e) and attempt < 5:
                print(f"    429 quota hit, backing off {delay}s...")
                time.sleep(delay)
                delay *= 2
                continue
            raise


def fetch_qcf(fitting_ids):
    """Per-fitting_id QC-fail journey timestamps from wms.order_items_history (IST).
    Returns {fitting_id: (IN_QC_0, QC_HOLD_0, QC_UNHOLD_0, QC_FAILED_0, IN_TRAY_1,
    EDGING_1, PENDING_CUSTOMIZATION_1, CUSTOMIZATION_COMPLETE_1, QC_DONE_1)}."""
    import pymysql
    cfg = dict(WMS_DB)
    cfg["connect_timeout"] = cfg.pop("connection_timeout", 15)
    cn = pymysql.connect(**cfg)
    try:
        with cn.cursor() as cur:
            placeholders = ",".join(["%s"] * len(fitting_ids))
            cur.execute(QCF_SQL.format(placeholders=placeholders), fitting_ids)
            rows = cur.fetchall()
    finally:
        cn.close()
    print(f"  wms.order_items_history: {len(rows)}/{len(fitting_ids)} fitting_ids matched")
    return {str(r[0]): r[1:] for r in rows}


def push_qcf(sh):
    """Write the wms journey timestamps straight into the QCF Analysis tab, joined to
    each row by fitting_id (col G). Fills M,N,O,P (QC IN/Hold/UN_Hold/Max QCF) and R
    (IN TRAY); REPLACES the existing T,V,X,Z formulas (min_mei_entry/mei_done/
    fitting_done/qc_done) with journey-1 values. Q and S are left untouched. Cells with
    no wms record are written blank (and counted). Backs up A:Z of the data rows first."""
    ws = sh.worksheet(QCF_TAB)

    g = ws.col_values(7)                                  # col G incl. header in [0]
    rows = [(i + 1, v.strip()) for i, v in enumerate(g) if i >= 1 and v.strip()]
    if not rows:
        print("  no fitting_ids in col G — nothing to do")
        return
    fids = [fid for _, fid in rows]
    rownums = [rn for rn, _ in rows]
    lo, hi = rownums[0], rownums[-1]
    if rownums != list(range(lo, hi + 1)):
        raise RuntimeError(f"col G rows are not contiguous ({lo}-{hi}); aborting")
    print(f"  {len(fids)} fitting_ids in col G (rows {lo}-{hi})")

    data = fetch_qcf(fids)                                # fails fast if wms unreachable

    backup = ws.get(f"A{lo}:Z{hi}")                       # only after the DB is reachable
    bpath = os.path.join(HERE, f"backup_{QCF_TAB.replace(' ', '_')}.csv")
    pd.DataFrame(backup).to_csv(bpath, index=False, header=False)
    print(f"  backed up A{lo}:Z{hi} -> {os.path.basename(bpath)}")

    def fmt(v):
        if v is None or v == "":
            return ""
        return v.strftime("%Y-%m-%d %H:%M:%S") if hasattr(v, "strftime") else str(v)

    blank_tuple = (None,) * len(QCF_COLS)
    batch = [{"range": f"{col}{lo}:{col}{hi}",
              "values": [[fmt(data.get(fid, blank_tuple)[ci])] for fid in fids]}
             for ci, col in enumerate(QCF_COLS)]
    _batch_update_with_retry(ws, batch)                  # USER_ENTERED -> dates parse

    matched = sum(1 for fid in fids if fid in data)
    print(f"  wrote {len(QCF_COLS)} cols ({','.join(QCF_COLS)}) x {len(fids)} rows to "
          f"'{QCF_TAB}' ({matched} matched, {len(fids) - matched} blank)")


def _trim(rows, key_col=None):
    """Drop trailing filler rows and trailing all-empty columns. A row is filler if
    key_col is set and that column is blank (CALC 2's formula grid), else if the whole
    row is blank. Interior blank rows are kept (RCA is a report layout with spacers)."""
    rows = [list(r) for r in rows]

    def filler(r):
        if key_col is not None:
            return len(r) <= key_col or (r[key_col] or "").strip() == ""
        return all((c or "").strip() == "" for c in r)

    while rows and filler(rows[-1]):
        rows.pop()
    if not rows:
        return rows
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]     # pad ragged rows
    while width > 1 and all((r[width - 1] or "").strip() == "" for r in rows):
        for r in rows:
            r.pop()
        width -= 1
    return rows


def _hex(c):
    """Sheets {red,green,blue} floats -> 'RRGGBB'."""
    return "".join(f"{int(round(c.get(k, 0) * 255)):02X}" for k in ("red", "green", "blue"))


def _fetch_format(svc, title, nrows, ncols):
    """Per-cell format {(row,col): {bg,bold,italic,size,fg,halign,valign,wrap}} and merges
    for the A1:<ncols><nrows> range of a tab."""
    from openpyxl.utils import get_column_letter
    if not nrows or not ncols:
        return {}, []
    rng = f"'{title}'!A1:{get_column_letter(ncols)}{nrows}"
    resp = svc.spreadsheets().get(
        spreadsheetId=SPREADSHEET_ID, ranges=[rng], includeGridData=True,
        fields="sheets(merges,data(rowData(values(effectiveFormat("
               "backgroundColor,horizontalAlignment,verticalAlignment,wrapStrategy,"
               "textFormat(bold,italic,fontSize,foregroundColor))))))").execute()
    sheet = resp["sheets"][0]
    merges = [(m.get("startRowIndex", 0), m.get("startColumnIndex", 0),
               m.get("endRowIndex", 0), m.get("endColumnIndex", 0))
              for m in sheet.get("merges", [])]
    H = {"LEFT": "left", "CENTER": "center", "RIGHT": "right"}
    V = {"TOP": "top", "MIDDLE": "center", "BOTTOM": "bottom"}
    fmt = {}
    for ri, row in enumerate(sheet.get("data", [{}])[0].get("rowData", [])):
        for ci, v in enumerate(row.get("values", [])):
            ef = v.get("effectiveFormat")
            if not ef:
                continue
            d = {}
            bg = ef.get("backgroundColor") or {}
            # API omits zero components, so default missing to 0 — else black {0,0,0} and
            # red {1,0,0} look like white {1,1,1} and get skipped. Skip only true white.
            if not all(bg.get(k, 0) >= 0.99 for k in ("red", "green", "blue")):
                d["bg"] = _hex(bg)
            tf = ef.get("textFormat", {})
            if tf.get("bold"):
                d["bold"] = True
            if tf.get("italic"):
                d["italic"] = True
            if tf.get("fontSize"):
                d["size"] = tf["fontSize"]
            fg = tf.get("foregroundColor") or {}
            if any(fg.get(k, 0) > 0.01 for k in ("red", "green", "blue")):
                d["fg"] = _hex(fg)                          # skip black/default
            if ef.get("horizontalAlignment") in H:
                d["halign"] = H[ef["horizontalAlignment"]]
            if ef.get("verticalAlignment") in V:
                d["valign"] = V[ef["verticalAlignment"]]
            if ef.get("wrapStrategy") == "WRAP":
                d["wrap"] = True
            if d:
                fmt[(ri, ci)] = d
    return fmt, merges


def export_excel(sh):
    """Copy EXPORT_TABS as static values into one .xlsx (formulas stripped), preserving
    background colours, bold and merged cells. Each sheet also gets wrap off, centered
    alignment and all-borders. Built in memory and uploaded to Google Drive under
    <DRCA_DRIVE_FOLDER>/<Year>/<MonthName>/. Returns the file's webViewLink."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    from openpyxl.utils import get_column_letter
    from googleapiclient.discovery import build

    svc = build("sheets", "v4", credentials=sh.client.auth)
    thin = Side(style="thin", color="000000")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    def fill_font(cell, f):
        """Apply source background / font (bold, italic, size, colour)."""
        if not f:
            return
        if f.get("bg"):
            cell.fill = PatternFill("solid", fgColor=f["bg"])
        fk = {k: f[k] for k in ("bold", "italic") if f.get(k)}
        if f.get("size"):
            fk["size"] = f["size"]
        if f.get("fg"):
            fk["color"] = f["fg"]
        if fk:
            cell.font = Font(**fk)

    def merge(ws, merges, maxr, maxc):
        for sr, sc, er, ec in merges:
            if er <= maxr and ec <= maxc and (er - sr) * (ec - sc) > 1:
                ws.merge_cells(start_row=sr + 1, start_column=sc + 1, end_row=er, end_column=ec)

    def autofit(ws, widths):
        for c, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(c + 1)].width = min(w, 50)

    wb = Workbook()
    wb.remove(wb.active)
    for src, dst, key_col, keep_format, as_is in EXPORT_TABS:
        ws = wb.create_sheet(title=dst)

        if as_is:
            # paste source exactly: its own values + colours + fonts + alignment + wrap +
            # merges, full width incl. blank spacer columns; thin borders only on A1:R37.
            vrows = sh.worksheet(src).get_all_values()
            nrows = len(vrows)
            ncols = max(max((len(r) for r in vrows), default=0), RCA_BORDER_COLS)
            fmt, merges = _fetch_format(svc, src, max(nrows, RCA_BORDER_ROWS), ncols)
            widths = [8] * ncols
            for r in range(1, nrows + 1):
                row = vrows[r - 1]
                for c in range(1, ncols + 1):
                    val = row[c - 1] if c - 1 < len(row) else ""
                    cell = ws.cell(row=r, column=c, value=val)
                    f = fmt.get((r - 1, c - 1))
                    fill_font(cell, f)
                    if f:
                        cell.alignment = Alignment(horizontal=f.get("halign"),
                                                   vertical=f.get("valign"),
                                                   wrap_text=bool(f.get("wrap")))
                    if val:
                        widths[c - 1] = max(widths[c - 1], len(val) + 2)
            autofit(ws, widths)
            for r in range(1, RCA_BORDER_ROWS + 1):        # thin borders A1:R37 only
                for c in range(1, RCA_BORDER_COLS + 1):
                    ws.cell(row=r, column=c).border = box
            merge(ws, merges, max(nrows, RCA_BORDER_ROWS), ncols)
            print(f"  {src} -> sheet '{dst}': {nrows} rows x {ncols} cols "
                  f"(as-is, borders A1:R{RCA_BORDER_ROWS}), {len(fmt)} fmt cells, {len(merges)} merges")
            continue

        # standard tabs: trimmed values, centered, all-borders, source bg/bold + merges
        rows = _trim(sh.worksheet(src).get_all_values(), key_col)   # displayed values, no formulas
        ncols = max((len(r) for r in rows), default=0)
        fmt, merges = _fetch_format(svc, src, len(rows), ncols) if keep_format else ({}, [])
        widths = [8] * ncols
        for r, row in enumerate(rows, 1):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = center
                cell.border = box
                fill_font(cell, fmt.get((r - 1, c - 1)))
                widths[c - 1] = max(widths[c - 1], len(val) + 2)
        autofit(ws, widths)
        merge(ws, merges, len(rows), ncols)
        print(f"  {src} -> sheet '{dst}': {len(rows)} rows x {ncols} cols"
              + (f", {len(fmt)} fmt cells, {len(merges)} merges" if keep_format else ""))

    import io
    d = pd.to_datetime(RCA_DATE).date()
    fname = drca_name()
    buf = io.BytesIO()
    wb.save(buf)
    link = upload_drca_to_drive(buf.getvalue(), fname, d.year, d.strftime("%B"))
    print(f"  uploaded '{fname}' to Drive {d.year}/{d.strftime('%B')}")
    if link:
        print(f"  {link}")
    return link


def build_dac(sh, date_str, out_path):
    """'View for Dispatch' workbook (DAC) saved to out_path. Two sheets:
    'CALC' = the CALC tab's cols A:L (header + rows where col L <> 0), and
    'RCA'  = the RCA tab's C25:H28 block."""
    from openpyxl import Workbook
    dt.date.fromisoformat(date_str)               # validate; raises ValueError on bad date

    def nonzero(v):
        v = (v or "").strip().replace(",", "")
        if v == "":
            return False
        try:
            return float(v) != 0
        except ValueError:
            return True                            # non-numeric text counts as <> 0

    calc = sh.worksheet("CALC").get_all_values()
    calc_rows = [r[:12] for i, r in enumerate(calc)
                 if i == 0 or (len(r) > 11 and nonzero(r[11]))]
    rca_rows = sh.worksheet("RCA").get("C25:H28")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "CALC"
    for row in calc_rows:
        ws1.append(row)
    ws2 = wb.create_sheet("RCA")
    for row in rca_rows:
        ws2.append(row)
    wb.save(out_path)


def main():
    mode = sys.argv[1].lower() if len(sys.argv) > 1 else "both"

    # Webapp helpers — take the RCA date as argv (the Next.js route shells out to these).
    if mode == "link":                            # python Push.py link <date>
        print(drca_drive_link(sys.argv[2]) or "NONE")
        return
    if mode == "dac":                             # python Push.py dac <date> <out.xlsx>
        build_dac(get_sheet(), sys.argv[2], sys.argv[3])
        print(f"  saved {sys.argv[3]}")
        return

    sh = get_sheet()
    print(f"Opened sheet: {sh.title}")

    if mode == "check":
        print("Tabs:", [ws.title for ws in sh.worksheets()])
        for t in (RAW_TAB, PACK_TAB, AWB_TAB):
            ws = sh.worksheet(t)
            print(f"  '{t}': {ws.row_count}x{ws.col_count} grid, "
                  f"{len(ws.get_all_values())} used rows")
        return

    if mode in ("both", "raw"):
        df = pd.read_csv(CSV_PATH, dtype=str).fillna("")
        print(f"Raw Power BI <- {CSV_PATH} ({len(df)} rows x {len(df.columns)} cols)")
        write_tab(sh, RAW_TAB, df)

    if mode in ("both", "packing"):
        print("Packing Data <- warehouse DB")
        try:
            write_tab(sh, PACK_TAB, fetch_packing_df())
        except Exception as e:
            print(f"  SKIPPED — cannot reach {DB['host']}:{DB['port']}? ({e})")
            print("  Connect to the warehouse network/VPN, then: python Push.py packing")

    if mode in ("both", "awb"):
        print("Dispatched AWBs <- ndd_shipments")
        try:
            normal, rescue = fetch_awbs()
            write_awbs(sh, normal, rescue)
        except Exception as e:
            print(f"  SKIPPED — cannot reach {DB['host']}:{DB['port']}? ({e})")
            print("  Connect to the warehouse network/VPN, then: python Push.py awb")

    if mode == "qcf":
        print("QCF Analysis <- wms.order_items_history (QC-fail journey timestamps)")
        try:
            push_qcf(sh)
        except Exception as e:
            print(f"  SKIPPED — cannot reach {WMS_DB['host']}:{WMS_DB['port']}? ({e})")
            print("  Connect to the warehouse network/VPN, then: python Push.py qcf")

    if mode == "excel":
        print(f"Export Excel <- {[t[0] for t in EXPORT_TABS]}")
        try:
            export_excel(sh)
        except RuntimeError as e:
            print(f"  {e}")


if __name__ == "__main__":
    main()
