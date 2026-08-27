"""Build a styled Decanting XLSX workbook from metadata and RFC 4180 CSV.

Usage:
    python decanting_xlsx.py OUTPUT.xlsx

Standard input must contain one compact JSON metadata object on the first line,
followed by an RFC 4180 CSV header and zero or more data rows.  The workbook is
written in openpyxl's write-only mode and atomically replaces OUTPUT.xlsx only
after the complete input has been validated and the workbook has been saved.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import math
import os
from pathlib import Path
import re
import sys
import tempfile
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DETAIL_SHEET = "Inventory vs DOH"
SUMMARY_SHEET = "Summary"

DEEP_BLUE = "1F4E79"
MEDIUM_BLUE = "305496"
TEAL = "006B6B"
PURPLE = "4B0082"
RED = "C00000"
GRN_TEAL = "0E7C7B"
BURGUNDY = "843C0C"
DEEP_ORANGE = "C55A11"
CALC_ORANGE = "ED7D31"
COMMENT_PURPLE = "7030A0"
DARK_GREY = "404040"

# Source colors are keyed rather than positional so shared sources retain the
# same color in Lens and Frame workbooks even when their column order differs.
SOURCE_COLORS: tuple[tuple[str, str], ...] = (
    ("putaway_pending", "9E480E"),
    ("egl_manual", "BF8F00"),
    ("pl_manual", RED),
    ("pl_10", DEEP_BLUE),
    ("pl_11", DARK_GREY),
    ("pl_40", "4472C4"),
    ("nxs1", "548235"),
    ("nxs2", COMMENT_PURPLE),
    ("asrs", "2E75B6"),
)

CALCULATION_HEADERS = {
    "total inventory",
    "nxs1 split",
    "7 day req",
    "7 day requirement",
    "7 day available",  # Legacy input, retained for exporter compatibility.
    "7 day shortage",
    "available other than asrs",
}
GRN_HEADERS = {"iqc status", "pid qty", "grn qty"}
COMMENT_HEADERS = {"decant comment grn pendncy", "decant comment", "comments"}
PID_HEADERS = {"pid", "product id", "parent product id"}
STATUS_HEADER_SUFFIXES = ("7 day status", "7day status", "10 day status", "10day status")

INTEGER_HEADERS = {
    "7 day doh",
    "10 day doh",
    "total inventory",
    "available other than asrs",
    "pid qty",
    "grn qty",
    "bulk required",
    "transfer pendency",
}
SOURCE_COUNT_HEADERS = {
    "asrs",
    "asrs inventory",
    "nxs1",
    "nxs1 inventory",
    "nxs2",
    "nxs2 inventory",
    "egl manual",
    "egl manual 05",
    "putaway pending",
    "pl manual",
    "pl 10",
    "pl 11",
    "pl 40",
}
DECIMAL_HEADERS = {
    "7 day req",
    "7 day requirement",
    "7 day available",
    "7 day shortage",
}

THIN_SIDE = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
HEADER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
NORMAL_FONT = Font(name="Arial", size=9)

SUFFICIENT_FILL = PatternFill("solid", fgColor="E2EFDA")
INSUFFICIENT_FILL = PatternFill("solid", fgColor="FFE0E0")
NO_STOCK_FILL = PatternFill("solid", fgColor="FFF2CC")
NEW_PID_FILL = PatternFill("solid", fgColor="FFE0E0")
SUFFICIENT_FONT = Font(name="Arial", color="375623", size=9, bold=True)
INSUFFICIENT_FONT = Font(name="Arial", color=RED, size=9, bold=True)
NO_STOCK_FONT = Font(name="Arial", color="7F6000", size=9, bold=True)
NEW_PID_FONT = Font(name="Arial", color=RED, size=9, bold=True)

SUMMARY_TITLE_FONT = Font(name="Arial", bold=True, size=13, color=DEEP_BLUE)
SUMMARY_LABEL_FONT = Font(name="Arial", bold=True, size=10)
SUMMARY_VALUE_FONT = Font(name="Arial", size=10)
SUMMARY_SECTION_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
SUMMARY_SECTION_FILL = PatternFill("solid", fgColor=DEEP_BLUE)
SUMMARY_WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
SUMMARY_WARNING_FONT = Font(name="Arial", color="7F6000", size=10)


def _canonical_header(value: str) -> str:
    """Return a separator-insensitive, lower-case column name."""
    return re.sub(r"[^a-z0-9]+", " ", value.strip().lower()).strip()


def _header_fill_color(header: str) -> str:
    canonical = _canonical_header(header)

    if canonical in COMMENT_HEADERS:
        return COMMENT_PURPLE
    if canonical in CALCULATION_HEADERS:
        return CALC_ORANGE
    if canonical in GRN_HEADERS:
        return GRN_TEAL
    if canonical == "flag":
        return RED
    if canonical == "7 day doh":
        return TEAL
    if canonical == "10 day doh":
        return PURPLE
    if canonical in {"bulk required", "increff required"}:
        return BURGUNDY
    if canonical == "transfer pendency":
        return DEEP_ORANGE
    if canonical in {"plc", "core new"}:
        return MEDIUM_BLUE

    machine_name = re.sub(r"[^a-z0-9]+", "_", header.strip().lower()).strip("_")
    for prefix, color in SOURCE_COLORS:
        if machine_name == prefix or machine_name.startswith(f"{prefix}_"):
            return color

    return DEEP_BLUE if canonical else DARK_GREY


def _column_width(header: str) -> float:
    canonical = _canonical_header(header)
    machine_name = re.sub(r"[^a-z0-9]+", "_", header.strip().lower()).strip("_")

    exact = {
        "pid": 18,
        "product id": 18,
        "parent product id": 18,
        "hsn classification": 22,
        "brand": 22,
        "product type": 22,
        "plc": 16,
        "core new": 16,
        "flag": 12,
        "7 day doh": 12,
        "10 day doh": 12,
        "ros units 7 day": 18,
        "ros day 7 day": 18,
        "ros day highest month": 23,
        "ros window start": 17,
        "ros window end": 17,
        "iqc status": 18,
        "pid qty": 12,
        "grn qty": 12,
        "transfer pendency": 18,
        "bulk required": 18,
        "total inventory": 16,
        "nxs1 split": 12,
        "7 day req": 14,
        "7 day requirement": 18,
        "7 day available": 16,
        "7 day shortage": 17,
        "available other than asrs": 24,
        "decant comment grn pendncy": 32,
        "decant comment": 32,
        "comments": 56,
    }
    if canonical in exact:
        return exact[canonical]
    if canonical.endswith(STATUS_HEADER_SUFFIXES):
        return 18
    if (
        canonical in SOURCE_COUNT_HEADERS
        or machine_name.endswith(("_count", "_doh_inv"))
        or canonical.endswith(" doh")
    ):
        return 14
    return 14


def _number_format(header: str) -> str | None:
    canonical = _canonical_header(header)
    machine_name = re.sub(r"[^a-z0-9]+", "_", header.strip().lower()).strip("_")

    if canonical in PID_HEADERS:
        return "@"
    if canonical == "nxs1 split":
        return "0.00%"
    if canonical.startswith("ros day"):
        return "0.0000"
    if canonical.startswith("ros units"):
        return "#,##0.####"
    if (
        canonical in DECIMAL_HEADERS
        or machine_name.endswith("_doh_inv")
        or (canonical.endswith(" doh") and canonical not in {"7 day doh", "10 day doh"})
    ):
        return "#,##0.00"
    if (
        canonical in INTEGER_HEADERS
        or canonical in SOURCE_COUNT_HEADERS
        or machine_name.endswith("_count")
        or machine_name.endswith("_qty")
    ):
        return "#,##0"
    return None


def _numeric_kind(header: str) -> str | None:
    number_format = _number_format(header)
    if number_format is None or number_format == "@":
        return None
    if number_format == "#,##0":
        return "integer"
    return "decimal"


def _parse_number(value: str, kind: str | None) -> str | int | float | None:
    if value == "":
        return None
    if kind is None:
        return value

    candidate = value.strip().replace(",", "")
    try:
        parsed = float(candidate)
    except ValueError:
        return value
    if not math.isfinite(parsed):
        return value
    if kind == "integer" and parsed.is_integer():
        return int(parsed)
    return parsed


def _safe_text_cell(cell: WriteOnlyCell, value: Any) -> None:
    """Force text into a string cell so a leading '=' cannot become a formula."""
    if isinstance(value, str):
        cell.data_type = "s"


def _require_text(metadata: Mapping[str, Any], key: str) -> str:
    value = metadata.get(key)
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"metadata.{key} must be a non-empty string")
    return value.strip()


def _require_nonnegative_integer(metadata: Mapping[str, Any], key: str) -> int:
    value = metadata.get(key)
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"metadata.{key} must be a non-negative integer")
    numeric = float(value)
    if not math.isfinite(numeric) or numeric < 0 or not numeric.is_integer():
        raise ValueError(f"metadata.{key} must be a non-negative integer")
    return int(numeric)


def _validate_count_mapping(value: Any, key: str) -> dict[str, int]:
    if not isinstance(value, dict):
        raise ValueError(f"metadata.{key} must be an object")
    result: dict[str, int] = {}
    for raw_name, raw_count in value.items():
        if not isinstance(raw_name, str) or not raw_name.strip():
            raise ValueError(f"metadata.{key} contains an invalid source name")
        if isinstance(raw_count, bool) or not isinstance(raw_count, (int, float)):
            raise ValueError(f"metadata.{key}.{raw_name} must be a non-negative integer")
        numeric = float(raw_count)
        if not math.isfinite(numeric) or numeric < 0 or not numeric.is_integer():
            raise ValueError(f"metadata.{key}.{raw_name} must be a non-negative integer")
        result[raw_name.strip()] = int(numeric)
    return result


def _validate_distribution(value: Any, key: str) -> list[tuple[str, int]]:
    if not isinstance(value, list):
        raise ValueError(f"metadata.{key} must be an array")
    result: list[tuple[str, int]] = []
    for index, item in enumerate(value):
        if not isinstance(item, dict):
            raise ValueError(f"metadata.{key}[{index}] must be an object")
        name = item.get("name")
        count = item.get("count")
        if not isinstance(name, str) or not name.strip():
            raise ValueError(f"metadata.{key}[{index}].name must be a non-empty string")
        if isinstance(count, bool) or not isinstance(count, (int, float)):
            raise ValueError(f"metadata.{key}[{index}].count must be a non-negative integer")
        numeric = float(count)
        if not math.isfinite(numeric) or numeric < 0 or not numeric.is_integer():
            raise ValueError(f"metadata.{key}[{index}].count must be a non-negative integer")
        result.append((name.strip(), int(numeric)))
    return result


def _validate_metadata(raw: Any) -> dict[str, Any]:
    if not isinstance(raw, dict):
        raise ValueError("the first input line must be a JSON object")

    title = _require_text(raw, "title")
    as_of_date = _require_text(raw, "asOfDate")
    window_start = _require_text(raw, "windowStart")
    window_end = _require_text(raw, "windowEnd")
    generated_at = _require_text(raw, "generatedAt")
    total_pids = _require_nonnegative_integer(raw, "totalPids")
    new_pids = _require_nonnegative_integer(raw, "newPids")
    grn_matched_pids = _require_nonnegative_integer(raw, "grnMatchedPids")
    if new_pids > total_pids:
        raise ValueError("metadata.newPids cannot exceed metadata.totalPids")
    if grn_matched_pids > total_pids:
        raise ValueError("metadata.grnMatchedPids cannot exceed metadata.totalPids")

    source_rows = _validate_count_mapping(raw.get("sourceRows"), "sourceRows")
    warnings = raw.get("warnings")
    if not isinstance(warnings, list) or any(not isinstance(item, str) for item in warnings):
        raise ValueError("metadata.warnings must be an array of strings")

    return {
        "title": title,
        "asOfDate": as_of_date,
        "windowStart": window_start,
        "windowEnd": window_end,
        "generatedAt": generated_at,
        "totalPids": total_pids,
        "newPids": new_pids,
        "grnMatchedPids": grn_matched_pids,
        "sourceRows": source_rows,
        "warnings": [warning.strip() for warning in warnings if warning.strip()],
        "decantDistribution": _validate_distribution(
            raw.get("decantDistribution"), "decantDistribution"
        ),
        "commentsDistribution": _validate_distribution(
            raw.get("commentsDistribution"), "commentsDistribution"
        ),
    }


def _display_key(value: str) -> str:
    special = {
        "powerBi": "Power BI rows",
        "scopedPids": "Scoped PIDs",
        "grn": "GRN rows",
        "inventory": "Inventory-matched PIDs",
        "products": "Product-matched PIDs",
        "googleSheets": "Google Sheets rows",
        "bigQuery": "BigQuery rows",
    }
    if value in special:
        return special[value]
    separated = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", value).replace("_", " ")
    return f"{separated.strip().title()} rows"


def _summary_rows(metadata: Mapping[str, Any]) -> list[tuple[str, Any, str]]:
    rows: list[tuple[str, Any, str]] = [
        (f"{metadata['title']} — Summary", None, "title"),
        ("", None, "blank"),
        ("As-of date", metadata["asOfDate"], "value"),
        (
            "ROS window",
            f"{metadata['windowStart']} through {metadata['windowEnd']}",
            "value",
        ),
        ("Generated at", metadata["generatedAt"], "value"),
        ("Total PIDs", metadata["totalPids"], "value"),
        ("Existing PIDs", metadata["totalPids"] - metadata["newPids"], "value"),
        ("New PIDs", metadata["newPids"], "value"),
        ("GRN-matched PIDs", metadata["grnMatchedPids"], "value"),
        ("", None, "blank"),
        ("Data Sources", None, "section"),
    ]
    rows.extend(
        (_display_key(name), count, "value")
        for name, count in metadata["sourceRows"].items()
    )

    warnings: Sequence[str] = metadata["warnings"]
    rows.extend((("", None, "blank"), ("Source Warnings", None, "section")))
    if warnings:
        rows.extend((f"Warning {index}", warning, "warning") for index, warning in enumerate(warnings, 1))
    else:
        rows.append(("Warnings", "None", "value"))

    rows.extend((("", None, "blank"), ("Decant Distribution", None, "section")))
    decant_distribution: Sequence[tuple[str, int]] = metadata["decantDistribution"]
    if decant_distribution:
        rows.extend((name, count, "value") for name, count in decant_distribution)
    else:
        rows.append(("No results", 0, "value"))

    rows.extend((("", None, "blank"), ("Comments Distribution", None, "section")))
    comments_distribution: Sequence[tuple[str, int]] = metadata["commentsDistribution"]
    if comments_distribution:
        rows.extend((name, count, "value") for name, count in comments_distribution)
    else:
        rows.append(("No results", 0, "value"))
    return rows


def _append_summary_sheet(workbook: Workbook, metadata: Mapping[str, Any]) -> None:
    sheet = workbook.create_sheet(SUMMARY_SHEET)
    sheet.column_dimensions["A"].width = 56
    sheet.column_dimensions["B"].width = 38

    for label, value, kind in _summary_rows(metadata):
        label_cell = WriteOnlyCell(sheet, value=label or None)
        value_cell = WriteOnlyCell(sheet, value=value)
        _safe_text_cell(label_cell, label_cell.value)
        _safe_text_cell(value_cell, value)

        if kind == "title":
            label_cell.font = SUMMARY_TITLE_FONT
        elif kind == "section":
            label_cell.font = SUMMARY_SECTION_FONT
            value_cell.font = SUMMARY_SECTION_FONT
            label_cell.fill = SUMMARY_SECTION_FILL
            value_cell.fill = SUMMARY_SECTION_FILL
        elif kind == "warning":
            label_cell.font = SUMMARY_LABEL_FONT
            value_cell.font = SUMMARY_WARNING_FONT
            label_cell.fill = SUMMARY_WARNING_FILL
            value_cell.fill = SUMMARY_WARNING_FILL
        elif kind != "blank":
            label_cell.font = SUMMARY_LABEL_FONT
            value_cell.font = SUMMARY_VALUE_FONT
            if isinstance(value, int):
                value_cell.number_format = "#,##0"
        sheet.append([label_cell, value_cell])


def _append_detail_sheet(
    workbook: Workbook,
    reader: Iterable[list[str]],
    headers: list[str],
) -> int:
    sheet = workbook.create_sheet(DETAIL_SHEET)
    sheet.freeze_panes = "D2"
    sheet.row_dimensions[1].height = 18

    for column_index, header in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = _column_width(header)

    header_cells: list[WriteOnlyCell] = []
    for header in headers:
        cell = WriteOnlyCell(sheet, value=header)
        cell.data_type = "s"
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=_header_fill_color(header))
        cell.alignment = HEADER_ALIGNMENT
        cell.border = CELL_BORDER
        header_cells.append(cell)
    sheet.append(header_cells)

    formats = [_number_format(header) for header in headers]
    numeric_kinds = [_numeric_kind(header) for header in headers]
    flag_columns = {
        index for index, header in enumerate(headers) if _canonical_header(header) == "flag"
    }
    status_columns = {
        index
        for index, header in enumerate(headers)
        if _canonical_header(header).endswith(STATUS_HEADER_SUFFIXES)
    }

    row_count = 0
    for csv_row in reader:
        if len(csv_row) != len(headers):
            raise ValueError(
                f"CSV row {row_count + 2} has {len(csv_row)} fields; expected {len(headers)}"
            )
        row_count += 1
        output_cells: list[WriteOnlyCell] = []
        for index, raw_value in enumerate(csv_row):
            value = _parse_number(raw_value, numeric_kinds[index])
            cell = WriteOnlyCell(sheet, value=value)
            _safe_text_cell(cell, value)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER_ALIGNMENT
            cell.border = CELL_BORDER
            if formats[index]:
                cell.number_format = formats[index]

            if index in status_columns:
                status = raw_value.strip().lower()
                if status == "sufficient":
                    cell.font = SUFFICIENT_FONT
                    cell.fill = SUFFICIENT_FILL
                elif status == "insufficient":
                    cell.font = INSUFFICIENT_FONT
                    cell.fill = INSUFFICIENT_FILL
                elif status in {"no stock", "nostock"}:
                    cell.font = NO_STOCK_FONT
                    cell.fill = NO_STOCK_FILL
            elif index in flag_columns and raw_value.strip().lower() == "new pid":
                cell.font = NEW_PID_FONT
                cell.fill = NEW_PID_FILL
            output_cells.append(cell)
        sheet.append(output_cells)

    last_column = get_column_letter(len(headers))
    last_row = max(1, row_count + 1)
    sheet.auto_filter.ref = f"A1:{last_column}{last_row}"
    return row_count


def _read_metadata_and_header(
    stream: io.TextIOBase,
) -> tuple[dict[str, Any], list[str], Iterable[list[str]]]:
    metadata_line = stream.readline()
    if not metadata_line:
        raise ValueError("stdin is empty; expected a JSON metadata line followed by CSV")
    try:
        raw_metadata = json.loads(metadata_line.lstrip("\ufeff"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"invalid metadata JSON on stdin line 1: {exc.msg}") from exc
    metadata = _validate_metadata(raw_metadata)

    reader = csv.reader(stream, dialect="excel", strict=True)
    try:
        headers = next(reader)
    except StopIteration as exc:
        raise ValueError("CSV header row is missing after the metadata line") from exc
    except csv.Error as exc:
        raise ValueError(f"invalid CSV header: {exc}") from exc

    if headers:
        headers[0] = headers[0].lstrip("\ufeff")
    if not headers or any(not header.strip() for header in headers):
        raise ValueError("CSV headers must be non-empty")
    if len(set(headers)) != len(headers):
        raise ValueError("CSV headers must be unique")
    return metadata, headers, reader


def build_workbook(output_path: Path, stream: io.TextIOBase) -> int:
    metadata, headers, reader = _read_metadata_and_header(stream)
    output_path = output_path.expanduser().resolve()
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError("output path must end with .xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists() and output_path.is_dir():
        raise ValueError(f"output path is a directory: {output_path}")

    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}.",
        suffix=".xlsx",
        dir=output_path.parent,
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook = Workbook(write_only=True)
        row_count = 0
        build_error: Exception | None = None
        try:
            row_count = _append_detail_sheet(workbook, reader, headers)
            if row_count != metadata["totalPids"]:
                build_error = ValueError(
                    "CSV row reconciliation failed: "
                    f"received {row_count}, metadata.totalPids is {metadata['totalPids']}"
                )
        except Exception as exc:
            # A write-only worksheet owns an open temporary XML stream. Saving
            # the partial workbook below closes that stream cleanly before the
            # generated XLSX is discarded, avoiding leaked openpyxl temp files.
            build_error = exc

        try:
            _append_summary_sheet(workbook, metadata)
            workbook.save(temporary_path)
        except Exception as save_error:
            if build_error is not None:
                raise build_error from save_error
            raise
        if build_error is not None:
            raise build_error

        # A quick structural read catches an incomplete ZIP before it can
        # replace a previously valid export.  read_only avoids loading cells.
        saved = load_workbook(temporary_path, read_only=True, data_only=True)
        try:
            if saved.sheetnames != [DETAIL_SHEET, SUMMARY_SHEET]:
                raise ValueError("saved workbook failed sheet reconciliation")
        finally:
            saved.close()
        os.replace(temporary_path, output_path)
        return row_count
    finally:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass


def _arguments(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a styled Lens/Frame Decanting XLSX from JSON metadata and CSV stdin."
    )
    parser.add_argument("output", type=Path, help="destination .xlsx path")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    arguments = _arguments(sys.argv[1:] if argv is None else argv)
    # newline='' is required for correct RFC 4180 parsing, including quoted
    # fields containing CRLF. utf-8-sig safely accepts an optional input BOM.
    stream = io.TextIOWrapper(sys.stdin.buffer, encoding="utf-8-sig", newline="")
    build_workbook(arguments.output, stream)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Decanting XLSX export failed: {exc}", file=sys.stderr)
        raise SystemExit(1)
